import json
import coverage
import requests
import os
import re
import subprocess
from openpyxl import Workbook
import time  # 引入time模块


def run(intention_provided ,current_time ,item):
    # Set the proxy port, or you won't be able to connect to the api
    os.environ["http_proxy"] = "127.0.0.1:7890"
    os.environ["https_proxy"] = "127.0.0.1:7890"

    # key = input("Please enter the valid key value：")
    key = "AIzaSyA5sC_73x_0zsbUfqZJnET_m3z1jsCF9hQ"
    url = f'https://generativelanguage.googleapis.com/v1/models/gemini-pro:generateContent?key={key}'
    headers = {'Content-Type': 'application/json'}

    # Create a new workbook
    workbook = Workbook()

    # Gets the default worksheet
    sheet = workbook.active

    # Set the header line
    sheet.cell(row=1, column=2, value="Is a Failure-inducing test case found?")
    sheet.cell(row=1, column=3, value="Test input")
    sheet.cell(row=1, column=4, value="Assertion")

    def remove_code_markers(input_text):
        # Use regular expressions to remove  "```" and "python"
        cleaned_text = re.sub(r'```|python|`', '', input_text)
        return cleaned_text

    # 处理输入的文本，移除每行中的特定标识符 "Test input:"，并清除行尾的空白
    def remove_test_input(text):
        lines = text.split('\n')
        cleaned_lines = []
        for line in lines:
            cleaned_line = line.replace('Test input:',
                                        '').strip()  # Remove the identifier and remove the trailing whitespace
            if cleaned_line:  # If the cleaned row is not empty, it is added to the list
                cleaned_lines.append(cleaned_line)
        return [cleaned_lines]  # Returns another list containing the processed text list

    # 计算文件中包含特定前缀 "Test input:" 的行数
    def count_test_inputs(filename):
        count = 0
        with open(filename, 'r') as file:
            for line in file:
                line = line.strip()
                if line.startswith('Test input:'):
                    count += 1
        return count

    # 从指定文件中读取并组织测试用例数据。该函数的工作流程是按照一定的格式将文件中的内容分组，每组代表一个测试用例
    def read_test_cases(filename):
        # Open the file for reading
        with open(filename, 'r') as file:
            content = file.read()

        # Use a regular expression to capture text between "Test input:" markers
        pattern = r"(?<=Test input:)(.*?)(?=Test input:|$)"
        matches = re.findall(pattern, content, re.DOTALL)

        # Initialize an empty list to store test cases
        test_cases = []

        # Process each match
        for match in matches:
            # Replace multiple consecutive newlines with a single newline
            cleaned_text = re.sub(r'\n+', '\n', match.strip())
            test_cases.append(cleaned_text)

        return test_cases

    def run_command_with_pipes(command, input_data):
        # Run the command and return its output as a result
        process = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE)
        return process.communicate(input_data.encode("utf-8"))[0]

    def ask_gemini(data):
        # 尝试请求的次数
        attempts = 5
        for attempt in range(attempts):
            try:
                # 创建一个请求会话
                with requests.Session() as s:
                    s.adapters.DEFAULT_RETRIES = 5  # 增加重连次数
                    s.keep_alive = False  # 关闭多余的持久连接
                    response = s.post(url, headers=headers, data=json.dumps(data), stream=True)
                    if response.status_code != 200:
                        raise ValueError("Failed to generate response: " + response.text)

                # 如果请求成功，跳出循环
                if response.ok:
                    return response
            except requests.exceptions.RequestException as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                # 如果不是最后一次尝试，则等待
                if attempt < attempts - 1:
                    time.sleep(60)  # 延时60秒
                else:
                    raise  # 最后一次尝试失败后抛出异常
    # 1、生成buggy程序意图文件
    while True:
        if intention_provided:
            break
        else:
            prompt_1 = "What is the intention of this program?"
            code = f'./Utils/DfPrompting/data/{current_time}/{item}/PUT_code.py'

            with open(code, encoding='UTF-8') as file:
                code_content = file.read()

            prompt = prompt_1 + "\n" + code_content
            # print(prompt)

            data = {
                "contents": [
                    {
                        "parts": [{"text": f"{prompt}"}]
                    }
                ]
            }
            print("Waiting for Gemini...")
            response = ask_gemini(data)
            # print(f"response status_code: {response.status_code}")
            if response.status_code == 500:
                print(response.status_code)
                continue
            else:

                try:
                    reply_content = json.loads(json.dumps(response.json(), indent=4, ensure_ascii=False))
                    ai_reply_intention = reply_content['candidates'][0]['content']['parts'][0]['text']
                    print("AI:", ai_reply_intention)

                    intention_file = f'./Utils/DfPrompting/data/{current_time}/{item}/intention.txt'
                    with open(intention_file, "w", encoding='UTF-8') as file:
                        file.write(ai_reply_intention)
                    break

                except Exception as e:
                    print(e)
                    continue

    # 2、生成两个参考程序
    i = 1
    while i <= 1:
        # print(f"Test {i}：")
        # generating reference versions
        # Prepare user input
        j = 1
        while j <= 2:
            user_input_1 = "Generate a python program based on the following intention.Only generate the program, not generate any explain."
            file_path_1 = f'./Utils/DfPrompting/data/{current_time}/{item}/intention.txt'
            with open(file_path_1, encoding='utf-8') as file:
                file_content_1 = file.read()

            combined_text_1 = user_input_1 + "\n" + file_content_1

            data = {
                "contents": [
                    {
                        "parts": [{"text": f"{combined_text_1}"}]
                    }
                ]
            }
            print("Waiting for Gemini...")
            response = ask_gemini(data)
            # print(f"response status_code: {response.status_code}")
            if response.status_code == 500:
                print(response.status_code)
                continue
            else:
                try:
                    reply_content = json.loads(json.dumps(response.json(), indent=4, ensure_ascii=False))
                    # print(reply_content)
                    ai_reply = reply_content['candidates'][0]['content']['parts'][0]['text']
                    print("AI:", ai_reply)
                    # Determine the directory
                    output_directory = f"./Utils/DfPrompting/data/{current_time}/{item}/test_{i}"
                    output_file_1 = f"{output_directory}/reference_version_{j}.py"

                    # Make sure the directory exists, and create it if it doesn't
                    os.makedirs(output_directory, exist_ok=True)

                    ai_reply = remove_code_markers(ai_reply)
                    # Writing the AI's response to a file
                    with open(output_file_1, 'w') as f:
                        f.write(ai_reply)

                    j += 1

                except Exception as e:
                    print(e)
                    continue

        print("Reference versions generating end!")

        coverages = []
        coverage_num = 0
        found = False
        coverage_found = False
        test_num = 1
        while test_num <= 10:
            print(f"Test inputs generation {test_num}:")
            # generating test inputs
            # Prepare user input
            user_input_2 = """Please generate diverse test inputs for the following code.Please put Test input: 
            as a separate line in front of each whole test input,no other explain is needed.
            note that the test function can be run directly with the whole test input to produce results.
            For example, the following example Python program, whose code is
            def gcd1 (a, b) :
                if b == 0 :
                    return a
                else :
                    return gcd1 (a, a%b)
            a = int(input())
            b = int(input())
            print(gcd1(a,b))
            Then the whose test input you need to give should be like :
            Test input:2
            3
            """
            file_path_2 = f"./Utils/DfPrompting/data/{current_time}/{item}/PUT_code.py"
            with open(file_path_2, encoding='utf-8') as file:
                file_content_2 = file.read()
            combined_text_2 = user_input_2 + "\n" + file_content_2

            data = {
                "contents": [
                    {
                        "parts": [{"text": f"{combined_text_2}"}]
                    }
                ]
            }
            print("Waiting for Gemini...")
            response = ask_gemini(data)

            if response.status_code == 500:
                print(response.status_code)
                continue
            else:
                try:
                    reply_content = json.loads(json.dumps(response.json(), indent=4, ensure_ascii=False))
                    # print(reply_content)
                    ai_reply = reply_content['candidates'][0]['content']['parts'][0]['text']
                    print("AI:", ai_reply)
                    output_file_2 = f"./Utils/DfPrompting/data/{current_time}/{item}/test_{i}/test_inputs_{test_num}.txt"
                    # Writing the AI's response to a file
                    with open(output_file_2, 'w') as f:
                        f.write(ai_reply)
                    test_num += 1
                except Exception as e:
                    print(e)
                    continue

            # Read the test case from the file
            test_cases = read_test_cases(output_file_2)
            num_test_cases = count_test_inputs(output_file_2)

            if len(test_cases) == 0:
                with open(output_file_2, encoding='utf-8') as file:
                    file_content = file.read()

                # Output after removing the Test input: identifier
                test_cases = remove_test_input(file_content)

            # print test cases
            for idx, test_case in enumerate(test_cases, start=1):
                # print(f"Test input {idx}:")
                # print(test_case)
                input_data = test_case

                print("input_data:")
                print(input_data)
                # print(input_data)
                if not input_data:
                    # print(f"Test input {idx} error!")
                    continue
                else:
                    # reference_version_1 running result
                    reference_version_1_file = f"./Utils/DfPrompting/data/{current_time}/{item}/test_{i}/reference_version_1.py"
                    # executable code
                    try:
                        result_1 = subprocess.run(["python", reference_version_1_file], input=input_data,
                                                  capture_output=True, text=True, timeout=10)
                        if result_1.returncode != 0:
                            continue
                    except subprocess.TimeoutExpired:
                        continue
                    result_1_output = result_1.stdout
                    # print("Output_1:")
                    # print(result_1_output)

                    # reference_version_2
                    reference_version_2_file = f"./Utils/DfPrompting/data/{current_time}/{item}/test_{i}/reference_version_2.py"
                    try:
                        # executable code
                        result_2 = subprocess.run(["python", reference_version_2_file], input=input_data,
                                                  capture_output=True, text=True, timeout=10)
                        if result_2.returncode != 0:
                            continue
                    except subprocess.TimeoutExpired:
                        continue

                    result_2_output = result_2.stdout
                    # print("Output_2:")
                    # print(result_2_output)

                    if result_1_output != result_2_output:
                        continue
                    else:
                        # code running result
                        code_file = f"./Utils/DfPrompting/data/{current_time}/{item}/PUT_code.py"
                        try:
                            # executable code
                            result_3 = subprocess.run(["python", code_file], input=input_data, capture_output=True,
                                                      text=True, timeout=10)

                        except subprocess.TimeoutExpired:
                            continue

                        PUT_output = result_3.stdout
                        # print("Output_code:")
                        # print(PUT_output)

                        if result_1_output == PUT_output:
                            # Define the command to run
                            command = ["coverage", "run", f"./Utils/DfPrompting/data/{current_time}/{item}/PUT_code.py"]

                            # Execute the command and get the output
                            output = run_command_with_pipes(command, input_data)

                            # Generate a coverage report
                            cov = coverage.Coverage()
                            cov.load()  # Load the coverage data

                            # Output coverage report to a file
                            cov_report_file = f"./Utils/DfPrompting/data/{current_time}/{item}/coverage_report.txt"
                            with open(cov_report_file, 'w') as report_file:
                                cov.report(file=report_file)

                            # Read the coverage report from the file
                            with open(cov_report_file, encoding='utf-8') as report_file:
                                coverage_report = report_file.read()

                            # Split coverage is reported in behavioral units
                            report_lines = coverage_report.split('\n')

                            # Initialize a variable to store coverage information of code.py
                            code_coverage = None

                            # Iterate over each line of the report
                            for line in report_lines:
                                # Look for the line that contains code.py
                                if 'code.py' in line:
                                    # Extract coverage information (assuming the coverage information is in the third column, separated by whitespace)
                                    code_coverage = line.split()[3]
                                    break
                            # print("code coverage:", code_coverage)
                            # Print coverage information for code.py
                            if code_coverage == '0':
                                code_coverage = float(code_coverage)
                            else:
                                code_coverage = float(code_coverage.strip('%'))
                            # print("code coverage:", code_coverage)

                            if coverage_num == 0:
                                coverages.append(code_coverage)
                                coverage_num += 1
                            else:
                                if code_coverage <= coverages[coverage_num - 1]:
                                    coverages.append(code_coverage)
                                    coverage_num += 1
                                    if coverage_num == 10:
                                        print("Branch coverage saturated.")
                                        coverage_found = True
                                        break
                                    else:
                                        continue
                                else:
                                    coverages = []
                                    coverage_num = 0
                                    coverages.append(code_coverage)
                                    coverage_num += 1
                                    continue

                        else:
                            row = i + 1
                            sheet.cell(row=row, column=1, value=i)
                            sheet.cell(row=row, column=2, value="Yes")
                            sheet.cell(row=row, column=3, value=input_data)
                            sheet.cell(row=row, column=4, value=result_1_output)
                            print(f"The result of test {i}:")
                            print("Failure-Inducing Test is")
                            print(input_data)
                            print()
                            print("The assert is")
                            print(result_1_output)
                            found = True
                            break

            if found:
                break

            if coverage_found:
                break

        if not found:
            row = i + 1
            sheet.cell(row=row, column=1, value=i)
            sheet.cell(row=row, column=2, value="NO")
            print(f"The result of test {i}:")
            print("No failure-inducing test is found.")

        i += 1

    file_path_coverage_1 = f"./Utils/DfPrompting/data/{current_time}/{item}/coverage_report.txt"
    if os.path.exists(file_path_coverage_1):
        os.unlink(file_path_coverage_1)

    file_path_coverage_2 = ".coverage"
    if os.path.exists(file_path_coverage_2):
        os.unlink(file_path_coverage_2)

    workbook.save(filename=f"./Utils/DfPrompting/data/{current_time}/{item}/results.xlsx")
    print("End!")
    if found:
        return found, input_data, result_1_output
    else:
        return found, "", ""

# if __name__ == '__main__':
#     intention_provided = input("Do you provide intention? Please enter True or False:")
#     run(intention_provided)