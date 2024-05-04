import json
from openpyxl import Workbook
from openpyxl.styles import Font


def export_excel_file(json_file_path, excel_file_path):
    # 读取JSON文件
    with open(json_file_path, 'r') as file:
        data = json.load(file)

    # 创建Excel工作簿和工作表
    workbook = Workbook()
    sheet = workbook.active

    # 设置表头
    headers = ["程序", "输入", "期望输出", "错误输出", "正确输出", "测试用例类型"]
    sheet.append(headers)

    # 填充数据
    for item in data:
        program = item.get('filename', '')

        # 处理数据，将 None 转换为 'null'
        input_data = "\n".join("null" if x is None else str(x) for x in item.get('input', []))
        expected_output = "\n".join("null" if x is None else str(x) for x in item.get('expectedOutput', []))
        if type(item.get('buggy_output')) is str:
            buggy_output = "\n".join("null" if x is None else str(x) for x in item.get('buggy_output', '').split("\n"))
            correct_output = "\n".join("null" if x is None else str(x) for x in item.get('correct_output', '').split("\n"))
        else:
            buggy_output = "\n".join("null" if x is None else str(x) for x in item.get('buggy_output', []))
            correct_output = "\n".join("null" if x is None else str(x) for x in item.get('correct_output', []))

        testcase_type = item.get('testcaseType', '')

        # 追加行到sheet
        row = [program, input_data, expected_output, buggy_output, correct_output, testcase_type]
        sheet.append(row)

    # 设置字体颜色
    for row in sheet.iter_rows(min_row=2, max_col=6, max_row=sheet.max_row):
        if row[5].value == "FT-IA":  # 检查测试用例类型是否为'FT-IA'
            for cell in row:
                cell.font = Font(color="FF0000")  # 设置字体颜色为红色

    # 保存Excel文件
    workbook.save(excel_file_path)
    print(f"Excel文件 '{excel_file_path}' 已成功创建并填充数据。")


# 使用示例
# current_time = "python_2024_04_16_18_30_40"
# json_path = f'data/{current_time}/TestcaseType.json'
# excel_path = f'data/{current_time}/analyzeResults.xlsx'
# export_excel_file(json_path, excel_path)
